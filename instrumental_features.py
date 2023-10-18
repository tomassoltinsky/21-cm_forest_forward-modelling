'''Functions to incorporate instrumental features

Version 18.10.2023

Spectrum smoothed by running boxcar and fixed boxcar
Gaussian white noise including Aeff/Tsys dependence on freq and intrinsic power-law radio spectrum
'''

import sys
import numpy as np
import matplotlib.pyplot as plt
import matplotlib as mpl
import scipy.signal as scisi
from astropy.convolution import convolve, Box1DKernel
from matplotlib.ticker import AutoMinorLocator
from matplotlib import gridspec
import matplotlib.patches as patches

#constants

lambda_0 = 21.11#rest wavelength of signal	cm
G   = 6.67e-8	#gravitational constant in	cm^3 g^-1 s^-2
m_p = 1.67e-24	#proton mass in				g
k_B = 1.38e-16	#boltzman constant in 		erg	K^-1
h_p = 6.63e-27	#Planck constant in			erg s
c   = 3.e10		#speed of light in			cm s^-1
Mpc = 3.086e24	#Megaparsec in				cm
mJy = 1e-29		#miliJansky in				J s^-1 m^-2

fsize = 16

def transF(optdep):
  return np.exp(-optdep)

def lambda_obs(redshift,velH):
  return lambda_0*(1+redshift)*(1-velH/c)	#Calculates observed shifted wavelength

def freq_obs(redshift,velH):
  return c/lambda_obs(redshift,velH)

def z_obs(redshift,velH):
  return lambda_obs(redshift,velH)/lambda_0-1.

import openpyxl

def excel_column(data_name,n_col):

    wb = openpyxl.load_workbook(data_name)
    sheet = wb.active
    N_rows = sheet.max_row-1

    col = np.empty(N_rows)

    for i_row in range(N_rows):
        col[i_row] = sheet.cell(row=i_row+2,column=n_col).value

    return col

def add_noise(frequency,dv,S_source,spec_index,t_integration,N_dish):
  
  #read uGMRT Aeff/Tsys from Fig. 8 in Braun et al. 2019
  freq_0 = excel_column('../observability/sensitivity/sens_uGMRT.xlsx',1)*1000
  ATsys_0 = excel_column('../observability/sensitivity/sens_uGMRT.xlsx',2)  

  index_freq = np.digitize(frequency,freq_0,right=True)	#Interpolate in the density (similar to redshift)

  freq_low  = freq_0[index_freq-1]
  freq_high = freq_0[index_freq]

  w_low = 1./(frequency-freq_low)/(1./(freq_high-frequency)+1./(frequency-freq_low))
  w_low = np.where(frequency==freq_low,1.,w_low)
  w_low = np.where(frequency==freq_high,0.,w_low)

  w_high = 1./(freq_high-frequency)/(1./(freq_high-frequency)+1./(frequency-freq_low))
  w_high = np.where(frequency==freq_high,1.,w_high)
  w_high = np.where(frequency==freq_low,0.,w_high)

  ATsys  = w_low*ATsys_0[index_freq-1]+w_high*ATsys_0[index_freq]

  #calculate standard deviation for telescope following equations 2-3 in Ciardi et al. 2015 MNRAS 453, 101-105; Datta et al. 2007 MNRAS 382, 809–818
  n_noise = np.sqrt(2)*k_B/1e7/ATsys/np.sqrt(N_dish*(N_dish-1)*t_integration*3600*dv*1000)/mJy

  #generate radio spectrum and normalize the noise
  S_v = S_source*np.power(frequency/147.,spec_index)
  sigma_noise = n_noise/S_v

  print('<Noise>       = %.3fmJy' % np.mean(n_noise))
  print('<sigma_noise> = %.5f' % np.mean(sigma_noise))

  #add random values from gaussian distribution to signal
  noise = np.random.normal(0.,sigma_noise,len(frequency))
  return noise
  


path = str(sys.argv[1])
z_name = float(sys.argv[2])
fX_name = float(sys.argv[3])
dvH = float(sys.argv[4])
telescope = str(sys.argv[5])
spec_res = float(sys.argv[6])
S_min_QSO = float(sys.argv[7])
alpha_R = float(sys.argv[8])
t_int = float(sys.argv[9])
LOS_ori = int(sys.argv[10])

Nlos = 100
filenum = int(np.floor(LOS_ori/Nlos))
LOS = LOS_ori-Nlos*filenum

datafile = str('%slos_regrid/200cMpclong_los_21cmREBIN_n%d_z%.3f_fX%.1f_dv%d_file%d.dat' % (path,Nlos,z_name,fX_name,dvH,filenum))
data  = np.fromfile(str(datafile),dtype=np.float32)
z     = data[0]	#redshift
Lbox  = data[5]/1e3	#box size
Nbins = int(data[7])					#Number of pixels/cells/bins in one line-of-sight
Nlos = int(data[8])						#Number of lines-of-sight
x_initial = 9
print(Nbins)
vel_axis = data[(x_initial+Nbins):(x_initial+2*Nbins)]#Hubble velocity along LoS in km/s
#lam   = lambda_obs(z,vel_axis*1.e5)
#freq  = freq_obs(z,vel_axis*1.e5)
redsh = z_obs(z,vel_axis*1e5)
freq = freq_obs(z,vel_axis*1e5)


z_min = redsh[-1]*0+5.72
z_max = redsh[0]

datafile = str('%stau/tau_200cMpclong_n%d_z%.3f_fX%.1f_dv%d_file%d.dat' %(path,Nlos,z_name,fX_name,dvH,filenum))
data = np.fromfile(str(datafile),dtype=np.float32)
tau = np.reshape(data,(Nlos,Nbins))[LOS]

signal_ori = transF(tau)

fig = plt.figure(figsize=(8.,4.75))
gs = gridspec.GridSpec(1,1)
ax0 = plt.subplot(gs[0,0])

ax0.plot(redsh,signal_ori,'-',color='black',label=r'$F_{\rm ori}$')
#ax0.legend(loc='lower left',frameon=False,fontsize=fsize,ncol=3)
ax0.set_xlim(z_min,z_max)
ax0.set_ylim(0.94,1.005)
ax0.set_yticks(np.arange(0.94,1.005,0.02))
ax0.set_xlabel(r'$z$', fontsize=fsize)
ax0.set_ylabel(r'$F=e^{-\tau_{21}}$', fontsize=fsize)
ax0.xaxis.set_minor_locator(AutoMinorLocator())
ax0.yaxis.set_minor_locator(AutoMinorLocator())
ax0.tick_params(axis='x',which='major',direction='in',bottom=True,top=True,left=True,right=True
		,length=10,width=1,labelsize=fsize)
ax0.tick_params(axis='y',which='major',direction='in',bottom=True,top=True,left=True,right=True
		,length=10,width=1,labelsize=fsize)
ax0.tick_params(axis='both',which='minor',direction='in',bottom=True,top=True,left=True,right=True
		,length=5,width=1)

#ax0.set_title(r'$\mathrm{log}f_{\rm X}=%.1f, S_{147\mathrm{MHz}}=%.1f\,\mathrm{mJy},\ \alpha_{\mathrm{R}}=%.2f,\ %s:\ \Delta\nu=%d\,\mathrm{kHz},\ t_{\mathrm{int}}=%d\mathrm{hr},\ \rm LOS-%d$' % (fX_name,S_min_QSO,alpha_R,telescope,spec_res,t_int,LOS_ori),fontsize=fsize)
#ax0.text(5.95,0.965,r'SNR=%.2f' % SNR,fontsize=fsize)

freq_min = c/lambda_0/1.e6/(1.+z_max)
freq_max = c/lambda_0/1.e6/(1.+z_min)
freq_label = np.round(np.arange(np.ceil(freq_min),np.floor(freq_max)+0.01,2.),0)
freq_posit = (c/lambda_0/1.e6/freq_label-1.-z_min)/(z_max-z_min)
ax0up = ax0.twiny()
ax0up.set_xlabel(r'$\nu_{obs}\ [\rm MHz]$', fontsize=fsize)
ax0up.set_xticks(freq_posit)
ax0up.set_xticklabels(freq_label)
ax0up.tick_params(axis='x',which='major',direction='in',bottom=False,top=True,left=True,right=True
		,length=10,width=1,labelsize=fsize)

freq_label = np.arange(np.ceil(freq_min*10)/10+0.4,np.floor(freq_max*10)/10+0.01,0.5)
freq_posit = (c/lambda_0/1.e6/freq_label-1.-z_min)/(z_max-z_min)
ax0upmin = ax0.twiny()
ax0upmin.set_xticks(freq_posit)
ax0upmin.tick_params(axis='x',which='major',direction='in',bottom=False,top=True,left=True,right=True
		,length=5,width=1,labeltop=False)

plt.tight_layout()
#plt.subplots_adjust(hspace=.0)
plt.savefig('method_present/spectrum_ori_%dcMpc_z%.1f_fX%s_LOS%d.png' % (Lbox,z,fX_name,LOS_ori))
plt.show()



v_pix = (freq[-1]-freq[0])/Nbins/1e3
Npix = int(np.round(spec_res/v_pix,0))
print('Convolve over '+str(Npix)+' pixels')
'''
box_kernel = Box1DKernel(Npix)
signal_smooth = convolve(signal_ori, box_kernel,boundary='fill', fill_value=1.)
ind_smooth = np.arange(0,Nbins+1,Npix)
redsh_smooth = redsh[ind_smooth]
signal_smooth = signal_smooth[ind_smooth]
'''
Nbins_smooth = int(np.floor(Nbins/Npix))
Nbins = Nbins_smooth*Npix
ind_smooth = np.arange(0,Nbins-1,Npix)+int(np.floor(Npix/2))
print(Nbins,len(ind_smooth))
redsh_smooth = redsh[ind_smooth]
freq_smooth = freq[ind_smooth]
signal_smooth = np.empty(Nbins_smooth)
for i in range(Nbins_smooth):
   signal_smooth[i] = np.mean(signal_ori[i*Npix:(i+1)*Npix])

print(Nbins_smooth,len(signal_smooth),len(signal_ori[(Nbins_smooth-1)*Npix:]))

fig = plt.figure(figsize=(8.,4.75))
gs = gridspec.GridSpec(1,1)
#print(redsh)
z_min = redsh[-1]*0+5.72
z_max = redsh[0]

v_21cm_z6 = c/lambda_0/7./1e6
S_21cm_z6 = S_min_QSO*np.power(v_21cm_z6/147.,alpha_R)
print('S_21cm(z=6.0) = %.2fmJy' % S_21cm_z6)
v_21cm_z6 = c/lambda_0/(z_min+1)/1e6
S_21cm_z6 = S_min_QSO*np.power(v_21cm_z6/147.,alpha_R)
print('S_21cm(z=%.1f) = %.2fmJy' % (z_min,S_21cm_z6))

ax0 = plt.subplot(gs[0,0])

#ax0.plot(redsh,signal_ori,'-',color='royalblue',label=r'$F_{\rm ori}$')
ax0.plot(redsh_smooth,signal_smooth,'-',color='black',label=r'$F_{\rm sig}$')
#ax0.legend(loc='lower left',frameon=False,fontsize=fsize,ncol=3)
ax0.set_xlim(z_min,z_max)
ax0.set_ylim(0.94,1.005)
ax0.set_yticks(np.arange(0.94,1.005,0.02))
ax0.set_xlabel(r'$z$', fontsize=fsize)
ax0.set_ylabel(r'$F=e^{-\tau_{21}}$', fontsize=fsize)
ax0.xaxis.set_minor_locator(AutoMinorLocator())
ax0.yaxis.set_minor_locator(AutoMinorLocator())
ax0.tick_params(axis='x',which='major',direction='in',bottom=True,top=True,left=True,right=True
		,length=10,width=1,labelsize=fsize)
ax0.tick_params(axis='y',which='major',direction='in',bottom=True,top=True,left=True,right=True
		,length=10,width=1,labelsize=fsize)
ax0.tick_params(axis='both',which='minor',direction='in',bottom=True,top=True,left=True,right=True
		,length=5,width=1)

#ax0.set_title(r'$\mathrm{log}f_{\rm X}=%.1f, S_{147\mathrm{MHz}}=%.1f\,\mathrm{mJy},\ \alpha_{\mathrm{R}}=%.2f,\ %s:\ \Delta\nu=%d\,\mathrm{kHz},\ t_{\mathrm{int}}=%d\mathrm{hr},\ \rm LOS-%d$' % (fX_name,S_min_QSO,alpha_R,telescope,spec_res,t_int,LOS_ori),fontsize=fsize)
#ax0.text(5.95,0.965,r'SNR=%.2f' % SNR,fontsize=fsize)

freq_min = c/lambda_0/1.e6/(1.+z_max)
freq_max = c/lambda_0/1.e6/(1.+z_min)
freq_label = np.round(np.arange(np.ceil(freq_min),np.floor(freq_max)+0.01,2.),0)
freq_posit = (c/lambda_0/1.e6/freq_label-1.-z_min)/(z_max-z_min)
ax0up = ax0.twiny()
ax0up.set_xlabel(r'$\nu_{obs}\ [\rm MHz]$', fontsize=fsize)
ax0up.set_xticks(freq_posit)
ax0up.set_xticklabels(freq_label)
ax0up.tick_params(axis='x',which='major',direction='in',bottom=False,top=True,left=True,right=True
		,length=10,width=1,labelsize=fsize)

freq_label = np.arange(np.ceil(freq_min*10)/10+0.4,np.floor(freq_max*10)/10+0.01,0.5)
freq_posit = (c/lambda_0/1.e6/freq_label-1.-z_min)/(z_max-z_min)
ax0upmin = ax0.twiny()
ax0upmin.set_xticks(freq_posit)
ax0upmin.tick_params(axis='x',which='major',direction='in',bottom=False,top=True,left=True,right=True
		,length=5,width=1,labeltop=False)

plt.tight_layout()
#plt.subplots_adjust(hspace=.0)
plt.savefig('method_present/spectrum_smooth_%dcMpc_z%.1f_fX%s_%s_%dkHz_LOS%d.png' % (Lbox,z,fX_name,telescope,spec_res,LOS_ori))
plt.show()



N_d = 30
signal_withnoise = signal_smooth+add_noise(freq_smooth/1e6,spec_res,S_min_QSO,alpha_R,t_int,N_d)

'''
prom = 1.001
tau_smooth = -np.log(signal_smooth)
tau_withnoise = -np.log(signal_withnoise)
tau_noise = -np.log(signal_withnoise-signal_smooth+1)

maxima = scisi.find_peaks(tau_smooth,prominence=tau_smooth*(1.-1./prom))
tau_signal = tau_smooth[maxima[0]]
maxima_signal = maxima[0]
maxima = scisi.find_peaks(tau_withnoise,prominence=tau_withnoise*(1.-1./prom))
tau_observed = tau_withnoise[maxima[0]]
maxima_observed = maxima[0]
maxima = scisi.find_peaks(tau_noise,prominence=tau_noise*(1.-1./prom))
tau_noise = tau_noise[maxima[0]]
maxima_noise = maxima[0]
'''

fig = plt.figure(figsize=(8.,4.75))
gs = gridspec.GridSpec(1,1)
#print(redsh)
z_min = redsh[-1]*0+5.72
z_max = redsh[0]

v_21cm_z6 = c/lambda_0/7./1e6
S_21cm_z6 = S_min_QSO*np.power(v_21cm_z6/147.,alpha_R)
print('S_21cm(z=6.0) = %.2fmJy' % S_21cm_z6)
v_21cm_z6 = c/lambda_0/(z_min+1)/1e6
S_21cm_z6 = S_min_QSO*np.power(v_21cm_z6/147.,alpha_R)
print('S_21cm(z=%.1f) = %.2fmJy' % (z_min,S_21cm_z6))

ax0 = plt.subplot(gs[0,0])

#ax0.plot(redsh,signal_ori,'-',color='royalblue',label=r'$F_{\rm ori}$')
ax0.plot(redsh_smooth,signal_withnoise,'-',color='black',label=r'$F_{\rm obs}$')
ax0.plot(redsh_smooth,signal_smooth,'--',color='darkorange',label=r'$F_{\rm sig}$')
ax0.legend(loc='lower left',frameon=False,fontsize=fsize,ncol=1)
ax0.set_xlim(z_min,z_max)
ax0.set_ylim(0.96,1.03)
ax0.set_yticks(np.arange(0.96,1.03,0.02))
ax0.set_xlabel(r'$z$', fontsize=fsize)
ax0.set_ylabel(r'$F=e^{-\tau_{21}}$', fontsize=fsize)
ax0.xaxis.set_minor_locator(AutoMinorLocator())
ax0.yaxis.set_minor_locator(AutoMinorLocator())
ax0.tick_params(axis='x',which='major',direction='in',bottom=True,top=True,left=True,right=True
		,length=10,width=1,labelsize=fsize)
ax0.tick_params(axis='y',which='major',direction='in',bottom=True,top=True,left=True,right=True
		,length=10,width=1,labelsize=fsize)
ax0.tick_params(axis='both',which='minor',direction='in',bottom=True,top=True,left=True,right=True
		,length=5,width=1)

ax0.set_title(r'$S_{147\mathrm{MHz}}=%.1f\,\mathrm{mJy},\ \alpha_{\mathrm{R}}=%.2f,\ t_{\mathrm{int}}=%d\mathrm{hr}$' % (S_min_QSO,alpha_R,t_int),fontsize=fsize)
#ax0.text(5.95,0.965,r'SNR=%.2f' % SNR,fontsize=fsize)

freq_min = c/lambda_0/1.e6/(1.+z_max)
freq_max = c/lambda_0/1.e6/(1.+z_min)
freq_label = np.round(np.arange(np.ceil(freq_min),np.floor(freq_max)+0.01,2.),0)
freq_posit = (c/lambda_0/1.e6/freq_label-1.-z_min)/(z_max-z_min)
ax0up = ax0.twiny()
ax0up.set_xlabel(r'$\nu_{obs}\ [\rm MHz]$', fontsize=fsize)
ax0up.set_xticks(freq_posit)
ax0up.set_xticklabels(freq_label)
ax0up.tick_params(axis='x',which='major',direction='in',bottom=False,top=True,left=True,right=True
		,length=10,width=1,labelsize=fsize)

freq_label = np.arange(np.ceil(freq_min*10)/10+0.4,np.floor(freq_max*10)/10+0.01,0.5)
freq_posit = (c/lambda_0/1.e6/freq_label-1.-z_min)/(z_max-z_min)
ax0upmin = ax0.twiny()
ax0upmin.set_xticks(freq_posit)
ax0upmin.tick_params(axis='x',which='major',direction='in',bottom=False,top=True,left=True,right=True
		,length=5,width=1,labeltop=False)

plt.tight_layout()
#plt.subplots_adjust(hspace=.0)
plt.savefig('method_present/spectrum_noisy_%dcMpc_z%.1f_fX%s_%s_%dkHz_Smin%.1fmJy_alphaR%.2f_t%dh_LOS%d.png' % (Lbox,z,fX_name,telescope,spec_res,S_min_QSO,alpha_R,t_int,LOS_ori))
plt.show()